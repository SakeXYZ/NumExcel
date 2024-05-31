from openpyxl import load_workbook

book = load_workbook("sheet.xlsx")
sheet = book.active
lists = []

for i in range(3, 42):
    try:
        one = sheet["F" + str(i)].value
        two = sheet["G" + str(i)].value
        if one is None or two is None:
            lists.append((' ', ' '))
        else:
            lists.append((int(one), int(two)))
    except TypeError as TypeErr:
        pass

test_book = load_workbook("test.xlsx")
test_sheet = test_book.active



count = 1
for i, k in lists:
    test_sheet['E' + str(count)] = count
    test_sheet['F' + str(count)] = str(i)
    test_sheet['G' + str(count)] = str(k)
    count += 1

test_book.save("test.xlsx")
test_book.close()
